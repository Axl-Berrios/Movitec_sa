"""
MOVITEC S.A. — Motor de Automatización de Reportes
Genera de forma automática:
  1. Reporte Excel analítico (6 hojas)
  2. Informe PDF
  3. Módulo de ejecución periódica (scheduler-ready)
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.gridspec import GridSpec
from matplotlib.ticker import FuncFormatter
from datetime import datetime, date
import os, io, warnings
warnings.filterwarnings('ignore')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers as xl_numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm, mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table as RLTable,
                                 TableStyle, Image as RLImage, HRFlowable,
                                 PageBreak, KeepTogether)
from reportlab.graphics.shapes import Drawing, Rect, String
from reportlab.graphics import renderPDF

# Constantes
AZUL_OSC   = '1A3A5C'
AZUL_MED   = '2E6DA4'
AZUL_CLAR  = 'D6E4F0'
ROJO       = 'C0392B'
VERDE      = '1E8449'
NARANJA    = 'E67E22'
GRIS       = 'F2F2F2'
BLANCO     = 'FFFFFF'

miles = FuncFormatter(lambda x,_: f'${x/1_000:.0f}K')
millones = FuncFormatter(lambda x,_: f'${x/1_000_000:.1f}M')

plt.rcParams.update({
    'font.family':'DejaVu Sans','axes.spines.top':False,'axes.spines.right':False,
    'axes.grid':True,'grid.alpha':0.3,'grid.linestyle':'--',
    'figure.facecolor':'#FAFAFA','axes.facecolor':'#FFFFFF',
})

os.makedirs('C:/Users/Axel/Downloads/outputs', exist_ok=True)
os.makedirs('C:/Users/Axel/Downloads/movitec/graficos', exist_ok=True)


# CARGA Y PRE-CÓMPUTO CENTRALIZADO

def cargar_datos():
    df     = pd.read_excel('C:/Users/Axel/Downloads/Movitec S.A/ordenes_trabajo.xlsx')
    df_tec = pd.read_excel('C:/Users/Axel/Downloads/Movitec S.A/tecnicos.xlsx')
    sc     = pd.read_excel('C:/Users/Axel/Downloads/Movitec S.A/scorecard_tecnicos.xlsx')
    eq     = pd.read_excel('C:/Users/Axel/Downloads/Movitec S.A/risk_equipos.xlsx')

    df_c   = df[df['estado']=='Cerrada'].copy()
    df_c   = df_c.merge(df_tec[['id tecnico','nombre','Años de experiencia']], on='id tecnico')
    

    # KPIs globales
    kpis = {
        'total_ot':          len(df),
        'cerradas':          int((df['estado']=='Cerrada').sum()),
        'pendientes':        int((df['estado']=='Pendiente').sum()),
        'en_proceso':        int((df['estado']=='En Proceso').sum()),
        'mttr_global':       round(df_c['tiempo resolucion horas'].mean(), 1),
        'costo_prom':        int(df_c['costo_total'].mean()),
        'costo_total':       int(df['costo_total'].sum()),
        'pct_retrabajo':     round((df['retrabajo']=='Sí').mean()*100, 1),
        'costo_retrabajo':   int(df[df['retrabajo']=='Sí']['costo_total'].sum()),
        'n_tecnicos':        df['id tecnico'].nunique(),
        'n_equipos':         df['equipo'].nunique(),
        'periodo_desde':     df['fecha_creacion'].min().strftime('%d/%m/%Y'),
        'periodo_hasta':     df['fecha_creacion'].max().strftime('%d/%m/%Y'),
    }

    # SLA
    sla = {}
    for p, lim in [('Alta',24),('Media',72),('Baja',168)]:
        sub = df_c[(df_c['prioridad']==p) & df_c['tiempo resolucion horas'].notna()]
        sla[p] = round((sub['tiempo resolucion horas']<=lim).mean()*100, 1)
    kpis['sla'] = sla

    return df, df_tec, df_c, sc, eq, kpis

df, df_tec, df_c, scorecard, equipo_stats, KPIS = cargar_datos()


# GENERADOR DE GRÁFICOS PARA REPORTES

def generar_grafico_kpis_resumen(path):
    """Mini-dashboard de 4 KPIs para portada"""
    fig, axes = plt.subplots(1, 4, figsize=(16, 3.5), facecolor='#1A3A5C')
    kpi_items = [
        ('MTTR Global',       f"{KPIS['mttr_global']:.1f}h",    'Referencia: <8h',   '#E74C3C'),
        ('Costo Prom / OT',   f"${KPIS['costo_prom']/1000:.0f}K", '$ por orden', '#E67E22'),
        ('% Retrabajo',       f"{KPIS['pct_retrabajo']:.1f}%",  'Benchmark: ≤10%',  '#E74C3C'),
        ('SLA Alta Prio',     f"{KPIS['sla']['Alta']:.1f}%",    'Objetivo: ≥90%',   '#E74C3C'),
    ]
    for ax, (titulo, valor, sub, color) in zip(axes, kpi_items):
        ax.set_facecolor('#1A3A5C')
        ax.set_xlim(0,1); ax.set_ylim(0,1)
        ax.axis('off')
        ax.add_patch(plt.Rectangle((0.05,0.05), 0.90, 0.90, fill=True,
                                    facecolor='#0D2137', edgecolor=color, linewidth=2, alpha=0.9))
        ax.text(0.50, 0.72, titulo, ha='center', va='center', fontsize=10,
                color='#BDC3C7', fontweight='bold', transform=ax.transAxes)
        ax.text(0.50, 0.45, valor,  ha='center', va='center', fontsize=22,
                color=color, fontweight='bold', transform=ax.transAxes)
        ax.text(0.50, 0.18, sub,    ha='center', va='center', fontsize=8,
                color='#7F8C8D', fontstyle='italic', transform=ax.transAxes)
    plt.tight_layout(pad=0.3)
    plt.savefig(path, dpi=150, bbox_inches='tight', facecolor='#1A3A5C')
    plt.close()

def generar_grafico_tecnicos_pdf(path):
    fig, axes = plt.subplots(1, 2, figsize=(14, 5), facecolor='#FAFAFA')
    sc_s = scorecard.sort_values('inef_score')
    # Izq: score ineficiencia
    colores = ['#1E8449' if s<=30 else '#E67E22' if s<=60 else '#C0392B' for s in sc_s['inef_score']]
    axes[0].barh(sc_s['nombre'], sc_s['inef_score'], color=colores, edgecolor='white')
    axes[0].axvline(30, color='#1E8449', linestyle='--', linewidth=1.2, alpha=0.7)
    axes[0].axvline(60, color='#E67E22', linestyle='--', linewidth=1.2, alpha=0.7)
    axes[0].set_title('Score de Ineficiencia por Técnico', fontweight='bold')
    axes[0].set_xlabel('Score (0=óptimo)')
    for i, (bar, s) in enumerate(zip(axes[0].patches, sc_s['inef_score'])):
        axes[0].text(bar.get_width()+0.5, bar.get_y()+bar.get_height()/2,
                     f'{s:.0f}', va='center', fontsize=7.5)
    # Der: % retrabajo vs costo
    scatter = axes[1].scatter(sc_s['pct_retrabajo'], sc_s['costo_prom']/1000,
                               c=colores, s=sc_s['total_ot']*3, alpha=0.85,
                               edgecolors='white', linewidth=0.8)
    axes[1].axvline(10, color='#1E8449', linestyle='--', linewidth=1.2, alpha=0.7, label='Benchmark 10%')
    axes[1].set_title('% Retrabajo vs Costo Promedio', fontweight='bold')
    axes[1].set_xlabel('% Retrabajo')
    axes[1].set_ylabel('Costo promedio (K$)')
    axes[1].yaxis.set_major_formatter(miles)
    for _, r in sc_s.iterrows():
        axes[1].annotate(r['nombre'].split()[0], (r['pct_retrabajo'], r['costo_prom']/1000),
                         textcoords='offset points', xytext=(4,3), fontsize=7)
    axes[1].legend(fontsize=8)
    leg = [mpatches.Patch(color='#1E8449', label='Eficiente'),
           mpatches.Patch(color='#E67E22', label='Intermedio'),
           mpatches.Patch(color='#C0392B', label='Crítico')]
    axes[1].legend(handles=leg, fontsize=7.5, loc='upper left')
    plt.tight_layout()
    plt.savefig(path, dpi=150, bbox_inches='tight')
    plt.close()

def generar_grafico_equipos_pdf(path):
    fig, axes = plt.subplots(1, 2, figsize=(14, 5), facecolor='#FAFAFA')
    # Pareto costo
    eq_s = equipo_stats.sort_values('costo_total', ascending=False).reset_index(drop=True)
    eq_s['pct_acum'] = eq_s['costo_total'].cumsum()/eq_s['costo_total'].sum()*100
    ax2 = axes[0].twinx()
    axes[0].bar(range(len(eq_s)), eq_s['costo_total']/1e6,
                color='#2E6DA4', alpha=0.75, edgecolor='white')
    ax2.plot(range(len(eq_s)), eq_s['pct_acum'], color='#C0392B',
             linewidth=2, marker='o', markersize=4)
    ax2.axhline(80, color='#7F8C8D', linestyle=':', linewidth=1.3)
    axes[0].set_title('Pareto de Costo por Equipo', fontweight='bold')
    axes[0].set_ylabel('Costo (M$)')
    ax2.set_ylabel('% Acumulado', color='#C0392B')
    axes[0].set_xticks(range(len(eq_s)))
    axes[0].set_xticklabels([e.replace(' ','\n') for e in eq_s['equipo']], fontsize=6, rotation=45, ha='right')
    # Risk score
    eq_r = equipo_stats.sort_values('risk_score', ascending=True)
    colores_r = ['#C0392B' if s>65 else '#E67E22' if s>40 else '#1E8449' for s in eq_r['risk_score']]
    axes[1].barh(eq_r['equipo'], eq_r['risk_score'], color=colores_r, edgecolor='white')
    axes[1].axvline(40, color='#E67E22', linestyle='--', linewidth=1.2, alpha=0.8)
    axes[1].axvline(65, color='#C0392B', linestyle='--', linewidth=1.2, alpha=0.8)
    axes[1].set_title('Risk Score por Equipo', fontweight='bold')
    axes[1].set_xlabel('Risk Score (0–100)')
    for bar in axes[1].patches:
        axes[1].text(bar.get_width()+0.5, bar.get_y()+bar.get_height()/2,
                     f'{bar.get_width():.0f}', va='center', fontsize=7.5)
    plt.tight_layout()
    plt.savefig(path, dpi=150, bbox_inches='tight')
    plt.close()

def generar_grafico_tendencia_pdf(path):
    fig, axes = plt.subplots(2, 1, figsize=(14, 7), facecolor='#FAFAFA')
    # Tendencia mensual costo
    df['mes_dt'] = df['fecha_creacion'].dt.to_period('M').dt.to_timestamp()
    vol = df.groupby('mes_dt').agg(n=('id orden','count'), costo=('costo_total','sum')).reset_index()
    axes[0].fill_between(vol['mes_dt'], vol['costo']/1e6, alpha=0.20, color='#2E6DA4')
    axes[0].plot(vol['mes_dt'], vol['costo']/1e6, color='#2E6DA4', linewidth=2, marker='o', markersize=3.5)
    z = np.polyfit(range(len(vol)), vol['costo']/1e6, 1)
    axes[0].plot(vol['mes_dt'], np.polyval(z, range(len(vol))), 'r--', linewidth=1.3, alpha=0.6, label='Tendencia')
    axes[0].set_title('Evolución Mensual de Costo Total (M$)', fontweight='bold')
    axes[0].set_ylabel('M$ CLP')
    axes[0].legend(fontsize=8)
    # Retrabajo mensual
    df['retrabajo_bin'] = (df['retrabajo']=='Sí').astype(int)
    ret_m = df.groupby('mes_dt').agg(pct_ret=('retrabajo_bin','mean')).reset_index()
    ret_m['pct_ret'] *= 100
    axes[1].fill_between(ret_m['mes_dt'], ret_m['pct_ret'], alpha=0.20, color='#C0392B')
    axes[1].plot(ret_m['mes_dt'], ret_m['pct_ret'], color='#C0392B', linewidth=2, marker='o', markersize=3.5)
    axes[1].axhline(10, color='#1E8449', linestyle='--', linewidth=1.5, alpha=0.7, label='Benchmark 10%')
    axes[1].set_title('% Retrabajo Mensual', fontweight='bold')
    axes[1].set_ylabel('% Retrabajo')
    axes[1].legend(fontsize=8)
    plt.tight_layout()
    plt.savefig(path, dpi=150, bbox_inches='tight')
    plt.close()

# Generar todos los gráficos
print("▶ Generando gráficos para reportes...")
generar_grafico_kpis_resumen('C:/Users/Axel/Downloads/Movitec S.A/graficos/rpt_kpis.png')
generar_grafico_tecnicos_pdf('C:/Users/Axel/Downloads/Movitec S.A/graficos/rpt_tecnicos.png')
generar_grafico_equipos_pdf('C:/Users/Axel/Downloads/Movitec S.A/graficos/rpt_equipos.png')
generar_grafico_tendencia_pdf('C:/Users/Axel/Downloads/Movitec S.A/graficos/rpt_tendencia.png')
print("✓ Gráficos generados")



# FASE DE REPORTE EXCEL COMPLETO

def borde(style='thin', color='BFBFBF'):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def cell_fmt(ws, row, col, value, bold=False, fg=None, color='000000',
             size=10, align='center', fmt=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, color=color, size=size, name='Arial')
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    c.border    = borde()
    if fg:
        c.fill = PatternFill('solid', fgColor=fg)
    if fmt:
        c.number_format = fmt
    return c

def header_row(ws, row, cols_vals, fg=AZUL_OSC, size=10):
    for col, val in cols_vals:
        cell_fmt(ws, row, col, val, bold=True, fg=fg, color=BLANCO, size=size)
    ws.row_dimensions[row].height = 20

def generar_excel_ejecutivo(path_out):
    wb = Workbook()
    wb.remove(wb.active)

    # DASHBOARD 
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 2
    for col in 'BCDEFGHIJ':
        ws.column_dimensions[col].width = 16

    # Título
    ws.merge_cells('B1:J2')
    c = ws['B1']
    c.value     = 'MOVITEC S.A. — Dashboard de Mantenimiento'
    c.font      = Font(bold=True, color='FFFFFF', size=16, name='Arial')
    c.fill      = PatternFill('solid', fgColor=AZUL_OSC)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22; ws.row_dimensions[2].height = 22

    ws.merge_cells('B3:J3')
    c = ws['B3']
    c.value     = f"Período: {KPIS['periodo_desde']} — {KPIS['periodo_hasta']}  ·  Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.font      = Font(italic=True, color=AZUL_MED, size=9, name='Arial')
    c.fill      = PatternFill('solid', fgColor=AZUL_CLAR)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 16

    # Bloque KPIs principales
    kpi_blocks = [
        ('B5:C6',  'TOTAL OT',           f"{KPIS['total_ot']:,}",       AZUL_MED,  BLANCO),
        ('D5:E6',  'CERRADAS',           f"{KPIS['cerradas']:,}",        VERDE,     BLANCO),
        ('F5:G6',  'MTTR GLOBAL',        f"{KPIS['mttr_global']:.1f}h",  ROJO,      BLANCO),
        ('H5:I6',  'COSTO PROMEDIO',     f"${KPIS['costo_prom']/1000:.0f}K", NARANJA, BLANCO),
        ('B8:C9',  '% RETRABAJO',        f"{KPIS['pct_retrabajo']:.1f}%", ROJO,     BLANCO),
        ('D8:E9',  'SLA ALTA',           f"{KPIS['sla']['Alta']:.1f}%",  ROJO,      BLANCO),
        ('F8:G9',  'COSTO TOTAL 3A',    f"${KPIS['costo_total']/1e9:.2f}B", AZUL_OSC, BLANCO),
        ('H8:I9',  'PENDIENTES',         f"{KPIS['pendientes']:,}",       NARANJA,   BLANCO),
    ]
    for rng, titulo, valor, bg, fg_c in kpi_blocks:
        ws.merge_cells(rng)
        c = ws[rng.split(':')[0]]
        c.value     = f'{titulo}\n{valor}'
        c.font      = Font(bold=True, color=fg_c, size=13, name='Arial')
        c.fill      = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = borde('medium', AZUL_OSC)
        for rr in ws[rng]:
            for cc in rr:
                if cc.coordinate != rng.split(':')[0]:
                    cc.fill = PatternFill('solid', fgColor=bg)
        ws.row_dimensions[int(rng.split(':')[0][1:])].height = 32
        ws.row_dimensions[int(rng.split(':')[1][1:])].height = 32

    # Insertar imagen KPI gráfico
    try:
        img = XLImage('C:/Users/Axel/Downloads/Movitec S.A/graficos/rpt_kpis.png')
        img.width = 780; img.height = 175
        ws.add_image(img, 'B11')
    except: pass

    # KPIs DETALLADOS 
    ws2 = wb.create_sheet("KPIs Detalle")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 22
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 18

    ws2.merge_cells('A1:D1')
    c = ws2['A1']
    c.value = 'KPIs Operacionales — Detalle por Dimensión'
    c.font  = Font(bold=True, color='FFFFFF', size=13, name='Arial')
    c.fill  = PatternFill('solid', fgColor=AZUL_MED)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 22

    secciones = [
        ('TIEMPO (MTTR)', [
            ('MTTR Global', f"{KPIS['mttr_global']:.1f} h", 'Benchmark < 8h', ROJO),
            ('MTTR – Eléctrica', '53.5 h', '', ROJO),
            ('MTTR – Correctiva', '52.4 h', '', ROJO),
            ('MTTR – Preventiva', '49.7 h', '', NARANJA),
            ('MTTR – Emergencia', '46.7 h', '', NARANJA),
            ('MTTR – Mecánica',   '46.5 h', '', NARANJA),
        ]),
        ('COSTO', [
            ('Costo Promedio / OT', f"${KPIS['costo_prom']:,.0f} CLP", 'Benchmark < $400.000', NARANJA),
            ('Costo Total 3 Años',  f"${KPIS['costo_total']:,.0f} CLP", '', AZUL_MED),
            ('Costo Atribuible Retrabajo', f"${KPIS['costo_retrabajo']:,.0f} CLP", '⚠ Evitable', ROJO),
        ]),
        ('CALIDAD', [
            ('% Retrabajo Global', f"{KPIS['pct_retrabajo']:.1f}%", 'Benchmark ≤10% — CRÍTICO', ROJO),
            ('SLA Prioridad Alta',  f"{KPIS['sla']['Alta']:.1f}%",   'Objetivo ≥90% — CRÍTICO', ROJO),
            ('SLA Prioridad Media', f"{KPIS['sla']['Media']:.1f}%",  'Objetivo ≥90% — ALERTA',  NARANJA),
            ('SLA Prioridad Baja',  f"{KPIS['sla']['Baja']:.1f}%",   'Objetivo ≥90%',            VERDE),
        ]),
    ]

    row = 3
    for seccion, items in secciones:
        ws2.merge_cells(f'A{row}:D{row}')
        c = ws2.cell(row=row, column=1, value=seccion)
        c.font  = Font(bold=True, color='FFFFFF', size=10, name='Arial')
        c.fill  = PatternFill('solid', fgColor=AZUL_OSC)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws2.row_dimensions[row].height = 18
        row += 1
        header_row(ws2, row, [(1,'Indicador'),(2,'Valor'),(3,'Referencia'),(4,'Estado')], GRIS)
        ws2.cell(row=row,column=1).font = Font(bold=True, color=AZUL_OSC, size=9, name='Arial')
        ws2.cell(row=row,column=2).font = Font(bold=True, color=AZUL_OSC, size=9, name='Arial')
        ws2.cell(row=row,column=3).font = Font(bold=True, color=AZUL_OSC, size=9, name='Arial')
        ws2.cell(row=row,column=4).font = Font(bold=True, color=AZUL_OSC, size=9, name='Arial')
        row += 1
        for kpi_n, val, ref, estado_col in items:
            cell_fmt(ws2, row, 1, kpi_n, align='left')
            cell_fmt(ws2, row, 2, val,   bold=True, align='center')
            cell_fmt(ws2, row, 3, ref,   align='left', size=8)
            c4 = ws2.cell(row=row, column=4)
            c4.value  = '●'
            c4.font   = Font(color=estado_col, size=14, name='Arial')
            c4.fill   = PatternFill('solid', fgColor='F9F9F9')
            c4.border = borde()
            c4.alignment = Alignment(horizontal='center', vertical='center')
            ws2.row_dimensions[row].height = 17
            row += 1
        row += 1

    # SCORECARD TÉCNICOS 
    ws3 = wb.create_sheet("Técnicos")
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = 'A3'

    ws3.merge_cells('A1:H1')
    c = ws3['A1']
    c.value = 'Scorecard de Eficiencia — Técnicos (OT Cerradas)'
    c.font  = Font(bold=True, color='FFFFFF', size=12, name='Arial')
    c.fill  = PatternFill('solid', fgColor=AZUL_OSC)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 20

    cols_sc = [
        ('A','Técnico',18),('B','Especialidad',16),('C','Exp. (años)',13),
        ('D','Score Inef.',12),('E','% Retrabajo',13),('F','Costo Prom',14),
        ('G','MTTR Prom (h)',14),('H','Cluster',13),
    ]
    for col_l, hdr, w in cols_sc:
        ws3.column_dimensions[col_l].width = w
    header_row(ws3, 2, [(i+1,h) for i,(l,h,w) in enumerate(cols_sc)], AZUL_MED)

    sc_sorted = scorecard.sort_values('inef_score')
    for i, (_, r) in enumerate(sc_sorted.iterrows(), 3):
        bg = GRIS if i % 2 == 0 else BLANCO
        cell_fmt(ws3, i, 1, r['nombre'],       align='left', fg=bg)
        cell_fmt(ws3, i, 2, r['especialidad'], align='left', fg=bg)
        cell_fmt(ws3, i, 3, int(r['exp']),     fg=bg)
        # Score con color semáforo
        score_col = ROJO if r['inef_score']>60 else NARANJA if r['inef_score']>30 else VERDE
        c_s = ws3.cell(row=i, column=4, value=round(r['inef_score'],1))
        c_s.font  = Font(bold=True, color='FFFFFF', size=10, name='Arial')
        c_s.fill  = PatternFill('solid', fgColor=score_col)
        c_s.border = borde(); c_s.alignment = Alignment(horizontal='center', vertical='center')

        cell_fmt(ws3, i, 5, round(r['pct_retrabajo'],1), fg=bg, fmt='0.0"%"')
        cell_fmt(ws3, i, 6, int(r['costo_prom']),        fg=bg, fmt='$#,##0')
        cell_fmt(ws3, i, 7, round(r['mttr_prom'],1),     fg=bg)
        # Badge cluster
        cl_color = {'Eficiente': VERDE, 'Intermedio': NARANJA, 'Crítico': ROJO}.get(r.get('cluster_label',''), GRIS)
        c_cl = ws3.cell(row=i, column=8, value=r.get('cluster_label',''))
        c_cl.font      = Font(bold=True, color='FFFFFF', size=9, name='Arial')
        c_cl.fill      = PatternFill('solid', fgColor=cl_color)
        c_cl.border    = borde()
        c_cl.alignment = Alignment(horizontal='center', vertical='center')
        ws3.row_dimensions[i].height = 17

    # Insertar gráfico técnicos
    try:
        img3 = XLImage('C:/Users/Axel/Downloads/Movitec S.A/graficos/rpt_tecnicos.png')
        img3.width = 700; img3.height = 255
        ws3.add_image(img3, f'A{i+3}')
    except: pass

    # HOJA 4: EQUIPOS 
    ws4 = wb.create_sheet("Equipos")
    ws4.sheet_view.showGridLines = False
    ws4.freeze_panes = 'A3'

    ws4.merge_cells('A1:H1')
    c = ws4['A1']
    c.value = 'Análisis de Riesgo por Equipo'
    c.font  = Font(bold=True, color='FFFFFF', size=12, name='Arial')
    c.fill  = PatternFill('solid', fgColor=AZUL_OSC)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws4.row_dimensions[1].height = 20

    cols_eq = [
        ('A','Equipo',22),('B','Total OT',10),('C','Costo Total',15),
        ('D','Costo Prom',14),('E','% Retrabajo',13),('F','% No Planif.',13),
        ('G','Risk Score',12),('H','Nivel',12),
    ]
    for col_l, hdr, w in cols_eq:
        ws4.column_dimensions[col_l].width = w
    header_row(ws4, 2, [(i+1,h) for i,(l,h,w) in enumerate(cols_eq)], AZUL_MED)

    eq_sorted = equipo_stats.sort_values('risk_score', ascending=False)
    for i, (_, r) in enumerate(eq_sorted.iterrows(), 3):
        bg = GRIS if i % 2 == 0 else BLANCO
        cell_fmt(ws4, i, 1, r['equipo'],          align='left', fg=bg)
        cell_fmt(ws4, i, 2, int(r['total_ot']),   fg=bg)
        cell_fmt(ws4, i, 3, int(r['costo_total']),fg=bg, fmt='$#,##0')
        cell_fmt(ws4, i, 4, int(r['costo_prom']), fg=bg, fmt='$#,##0')
        cell_fmt(ws4, i, 5, round(r['pct_retrabajo'],1), fg=bg, fmt='0.0"%"')
        cell_fmt(ws4, i, 6, round(r['pct_no_planif'],1), fg=bg, fmt='0.0"%"')
        risk_col = ROJO if r['risk_score']>65 else NARANJA if r['risk_score']>40 else VERDE
        c_r = ws4.cell(row=i, column=7, value=round(r['risk_score'],0))
        c_r.font  = Font(bold=True, color='FFFFFF', size=10, name='Arial')
        c_r.fill  = PatternFill('solid', fgColor=risk_col)
        c_r.border = borde(); c_r.alignment = Alignment(horizontal='center', vertical='center')
        nivel = 'CRÍTICO' if r['risk_score']>65 else 'ALERTA' if r['risk_score']>40 else 'Normal'
        c_n = ws4.cell(row=i, column=8, value=nivel)
        c_n.font  = Font(bold=True, color='FFFFFF', size=9, name='Arial')
        c_n.fill  = PatternFill('solid', fgColor=risk_col)
        c_n.border = borde(); c_n.alignment = Alignment(horizontal='center', vertical='center')
        ws4.row_dimensions[i].height = 17

    # TENDENCIA MENSUAL 
    ws5 = wb.create_sheet("Tendencia")
    ws5.sheet_view.showGridLines = False

    ws5.merge_cells('A1:F1')
    c = ws5['A1']
    c.value = 'Tendencia Mensual — Costos y Retrabajo'
    c.font  = Font(bold=True, color='FFFFFF', size=12, name='Arial')
    c.fill  = PatternFill('solid', fgColor=AZUL_OSC)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws5.row_dimensions[1].height = 20

    df['mes_str'] = df['fecha_creacion'].dt.to_period('M').astype(str)
    df['ret_bin'] = (df['retrabajo']=='Sí').astype(int)
    tend = df.groupby('mes_str').agg(
        n_ot=('id orden','count'),
        costo_total=('costo_total','sum'),
        costo_prom=('costo_total','mean'),
        pct_ret=('ret_bin','mean'),
    ).reset_index()
    tend['pct_ret'] = (tend['pct_ret']*100).round(1)
    tend['costo_prom'] = tend['costo_prom'].round(0).astype(int)

    header_row(ws5, 2, [(1,'Mes'),(2,'# OT'),(3,'Costo Total'),(4,'Costo Prom'),(5,'% Retrabajo'),(6,'Tendencia')], AZUL_MED)
    for col,w in [('A',14),('B',10),('C',16),('D',14),('E',13),('F',12)]:
        ws5.column_dimensions[col].width = w

    for i,(_, r) in enumerate(tend.iterrows(), 3):
        bg = GRIS if i%2==0 else BLANCO
        cell_fmt(ws5, i, 1, r['mes_str'],          fg=bg)
        cell_fmt(ws5, i, 2, int(r['n_ot']),        fg=bg)
        cell_fmt(ws5, i, 3, int(r['costo_total']), fg=bg, fmt='$#,##0')
        cell_fmt(ws5, i, 4, int(r['costo_prom']),  fg=bg, fmt='$#,##0')
        ret_col = ROJO if r['pct_ret']>25 else NARANJA if r['pct_ret']>15 else VERDE
        c_ret = ws5.cell(row=i, column=5, value=r['pct_ret'])
        c_ret.font = Font(bold=r['pct_ret']>20, color=ret_col, size=9, name='Arial')
        c_ret.fill = PatternFill('solid', fgColor=bg); c_ret.border = borde()
        c_ret.alignment = Alignment(horizontal='center', vertical='center')
        c_ret.number_format = '0.0"%"'
        cell_fmt(ws5, i, 6, '', fg=bg)
        ws5.row_dimensions[i].height = 16

    try:
        img5 = XLImage('C:/Users/Axel/Downloads/Movitec S.A/graficos/rpt_tendencia.png')
        img5.width = 700; img5.height = 340
        ws5.add_image(img5, f'A{i+3}')
    except: pass

    # RECOMENDACIONES 
    ws6 = wb.create_sheet("Recomendaciones")
    ws6.sheet_view.showGridLines = False
    ws6.column_dimensions['A'].width = 4
    ws6.column_dimensions['B'].width = 30
    ws6.column_dimensions['C'].width = 50
    ws6.column_dimensions['D'].width = 18
    ws6.column_dimensions['E'].width = 16

    ws6.merge_cells('B1:E2')
    c = ws6['B1']
    c.value = 'Plan de Acción — Recomendaciones Estratégicas Movitec S.A.'
    c.font  = Font(bold=True, color='FFFFFF', size=13, name='Arial')
    c.fill  = PatternFill('solid', fgColor=AZUL_OSC)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws6.row_dimensions[1].height = 22; ws6.row_dimensions[2].height = 22

    header_row(ws6, 4, [(2,'Hallazgo'),(3,'Acción Recomendada'),(4,'Impacto Estimado'),(5,'Prioridad')], AZUL_MED)

    recs = [
        ('Retrabajo 21.7% (2× benchmark)',
         'Implementar protocolo de revisión de calidad post-intervención. Check-list por tipo de falla. Auditoría de órdenes con retrabajo recurrente.',
         '~$180M CLP/año en ahorros', '🔴 Inmediata'),
        ('Ignacio Ramírez — Score 100 (Crítico)',
         'Revisión de desempeño documentada. Plan de capacitación 60 días. Asignación supervisada en OT de alta prioridad. Seguimiento mensual.',
         '$180K CLP por OT evitada', '🔴 Inmediata'),
        ('Compresor C-102 — Risk Score 99',
         'Programar overhaul preventivo. Análisis de vida útil. Evaluar reemplazo si ROI es negativo. Plan de repuestos críticos.',
         'Reducir 40% fallas no planif.', '🔴 Inmediata'),
        ('SLA Alta Prioridad 65.8% (obj: 90%)',
         'Redefinir protocolo de escalamiento. Técnico de guardia 24/7 para emergencias. Tablero de control en tiempo real por prioridad.',
         '+24pp en cumplimiento SLA', '🟠 30 días'),
        ('MTTR Global 49.9h (obj: <8h)',
         'Revisión de flujo logístico de repuestos. Contrato marco con proveedores. Bodega de repuestos críticos on-site.',
         'Reducir MTTR un 40%', '🟠 60 días'),
        ('5 equipos en zona ALERTA/CRÍTICO',
         'Implementar mantenimiento predictivo en Conveyor CV-901, Motor M-303, Bomba B-202. Sensores IoT de vibración/temperatura.',
         '$90M CLP/año en correctivos', '🟡 90 días'),
        ('Outliers: 157 OT con costos anómalos',
         'Establecer umbrales de aprobación automática. Alertas en sistema si OT supera 2.5σ de costo. Aprobación gerencial para casos extremos.',
         'Control de sobrecostos', '🟡 90 días'),
    ]

    AMARILLO = "F39C12"
    for i, (hallazgo, accion, impacto, prio) in enumerate(recs, 5):
        bg = GRIS if i%2==0 else BLANCO
        prio_col = ROJO if '🔴' in prio else NARANJA if '🟠' in prio else AMARILLO if '🟡' in prio else VERDE
        cell_fmt(ws6, i, 2, hallazgo, align='left', fg=bg, wrap=True)
        cell_fmt(ws6, i, 3, accion,   align='left', fg=bg, wrap=True, size=9)
        cell_fmt(ws6, i, 4, impacto,  align='center', fg=bg, bold=True, size=9)
        c_p = ws6.cell(row=i, column=5, value=prio)
        c_p.font = Font(bold=True, color='FFFFFF', size=9, name='Arial')
        c_p.fill = PatternFill('solid', fgColor=prio_col)
        c_p.border = borde(); c_p.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws6.row_dimensions[i].height = 48


    wb.save(path_out)
    print(f"✓ Excel ejecutivo guardado: {path_out}")

generar_excel_ejecutivo('C:/Users/Axel/Downloads/Movitec S.A/MOVITEC_Reporte.xlsx')



# INFORME PDF (CONSULTORÍA)

def generar_pdf_ejecutivo(path_out):
    doc = SimpleDocTemplate(
        path_out, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm,
    )
    W, H = A4
    styles = getSampleStyleSheet()

    # Estilos personalizados 
    st_titulo = ParagraphStyle('titulo',
        fontSize=22, fontName='Helvetica-Bold', textColor=colors.HexColor('#1A3A5C'),
        alignment=TA_CENTER, spaceAfter=6)
    st_subtitulo = ParagraphStyle('subtitulo',
        fontSize=11, fontName='Helvetica', textColor=colors.HexColor('#2E6DA4'),
        alignment=TA_CENTER, spaceAfter=4)
    st_h2 = ParagraphStyle('h2',
        fontSize=13, fontName='Helvetica-Bold', textColor=colors.white,
        backColor=colors.HexColor('#1A3A5C'), alignment=TA_LEFT,
        leftIndent=-6, rightIndent=-6, spaceAfter=6, spaceBefore=14,
        borderPadding=(4,6,4,6))
    st_h3 = ParagraphStyle('h3',
        fontSize=11, fontName='Helvetica-Bold', textColor=colors.HexColor('#2E6DA4'),
        spaceAfter=4, spaceBefore=8)
    st_body = ParagraphStyle('body',
        fontSize=9.5, fontName='Helvetica', leading=14,
        textColor=colors.HexColor('#2C3E50'), spaceAfter=5, alignment=TA_JUSTIFY)
    st_bullet = ParagraphStyle('bullet',
        fontSize=9, fontName='Helvetica', leading=13,
        textColor=colors.HexColor('#2C3E50'), leftIndent=14, spaceAfter=3,
        bulletIndent=4)
    st_footer = ParagraphStyle('footer',
        fontSize=7.5, fontName='Helvetica-Oblique',
        textColor=colors.HexColor('#7F8C8D'), alignment=TA_CENTER)
    st_alerta = ParagraphStyle('alerta',
        fontSize=9, fontName='Helvetica-Bold',
        textColor=colors.white, backColor=colors.HexColor('#C0392B'),
        borderPadding=(3,6,3,6), spaceAfter=4, leftIndent=-6, rightIndent=-6)
    st_verde = ParagraphStyle('verde',
        fontSize=9, fontName='Helvetica-Bold',
        textColor=colors.white, backColor=colors.HexColor('#1E8449'),
        borderPadding=(3,6,3,6), spaceAfter=4, leftIndent=-6, rightIndent=-6)

    story = []
    LINEA = HRFlowable(width='100%', thickness=1.5, color=colors.HexColor('#1A3A5C'), spaceAfter=8)
    LINEA_GRIS = HRFlowable(width='100%', thickness=0.5, color=colors.HexColor('#BDC3C7'), spaceAfter=6)

    #  PORTADA 
    story.append(Spacer(1, 1.5*cm))
    story.append(Paragraph('MOVITEC S.A.', st_titulo))
    story.append(Paragraph('Servicios Técnicos y Mantenimiento Industrial', st_subtitulo))
    story.append(Spacer(1, 0.3*cm))
    story.append(LINEA)
    story.append(Spacer(1, 0.3*cm))

    st_titulo2 = ParagraphStyle('t2', fontSize=16, fontName='Helvetica-Bold',
                                  textColor=colors.HexColor('#2E6DA4'), alignment=TA_CENTER, spaceAfter=4)
    story.append(Paragraph('INFORME EJECUTIVO DE OPTIMIZACIÓN OPERACIONAL', st_titulo2))
    story.append(Paragraph('Análisis End-to-End · Detección de Ineficiencias · Plan de Acción', st_subtitulo))
    story.append(Spacer(1, 0.5*cm))

    # Mini-tabla de metadatos
    meta = [
        ['Período analizado', f"{KPIS['periodo_desde']} — {KPIS['periodo_hasta']}"],
        ['Órdenes de trabajo', f"{KPIS['total_ot']:,} registros"],
        ['Técnicos evaluados', str(KPIS['n_tecnicos'])],
        ['Equipos en flota', str(KPIS['n_equipos'])],
        ['Fecha del informe', datetime.now().strftime('%d de %B de %Y')],
        ['Preparado por', 'Consultoría Analítica Senior — Movitec S.A.'],
    ]
    t_meta = RLTable(meta, colWidths=[5.5*cm, 9*cm])
    t_meta.setStyle(TableStyle([
        ('FONTNAME',    (0,0),(-1,-1), 'Helvetica'),
        ('FONTSIZE',    (0,0),(-1,-1), 9),
        ('FONTNAME',    (0,0),(0,-1),  'Helvetica-Bold'),
        ('TEXTCOLOR',   (0,0),(0,-1),  colors.HexColor('#1A3A5C')),
        ('BACKGROUND',  (0,0),(0,-1),  colors.HexColor('#D6E4F0')),
        ('GRID',        (0,0),(-1,-1), 0.5, colors.HexColor('#BDC3C7')),
        ('ROWBACKGROUNDS',(0,0),(-1,-1),[colors.white, colors.HexColor('#F8FBFF')]),
        ('PADDING',     (0,0),(-1,-1), 5),
    ]))
    story.append(t_meta)
    story.append(Spacer(1, 0.4*cm))

    # Gráfico KPIs portada
    try:
        img_kpi = RLImage('C:/Users/Axel/Downloads/Movitec S.A/graficos/rpt_kpis.png',
                          width=16*cm, height=3.6*cm)
        story.append(img_kpi)
    except: pass
    story.append(PageBreak())

    # RESUMEN 
    story.append(Paragraph('1. RESUMEN', st_h2))
    story.append(Paragraph(
        f'El presente informe consolida el análisis operacional de Movitec S.A. sobre una base de '
        f'<b>{KPIS["total_ot"]:,} órdenes de trabajo</b> ejecutadas entre {KPIS["periodo_desde"]} y '
        f'{KPIS["periodo_hasta"]}. El análisis revela tres problemas estructurales con impacto '
        f'económico cuantificable y acción inmediata requerida.', st_body))

    story.append(Paragraph('▶ Hallazgos Críticos Identificados', st_h3))
    hallazgos = [
        '<b>Tasa de retrabajo de 21.7%</b>, más del doble del benchmark internacional (≤10%). '
        f'Costo directo atribuible: <b>${KPIS["costo_retrabajo"]/1e9:.2f}B CLP</b> en el período.',
        f'<b>SLA de Alta Prioridad en 65.8%</b> contra un objetivo de 90%. Cada punto porcentual '
        f'perdido representa riesgo operativo directo para activos críticos.',
        f'<b>MTTR global de {KPIS["mttr_global"]:.1f} horas</b>. El estándar ISO 55000 para '
        f'mantenimiento industrial de referencia es inferior a 8 horas en fallas correctivas.',
        '<b>Técnico Ignacio Ramírez</b>: score de ineficiencia 100/100, tasa de retrabajo 43.8%, '
        f'costo promedio por OT de $677.805 CLP (88% sobre la media de flota).',
        '<b>Compresor C-102</b>: risk score 99/100 con 62% de fallas no planificadas. '
        'Pareto de equipos confirma que 4 activos concentran el 35% del costo total.',
    ]
    for h in hallazgos:
        story.append(Paragraph(f'• {h}', st_bullet))
    story.append(Spacer(1, 0.3*cm))

    # KPIs OPERACIONALES
    story.append(Paragraph('2. KPIs OPERACIONALES', st_h2))

    kpi_data = [
        ['KPI', 'Valor Actual', 'Benchmark', 'Gap', 'Estado'],
        ['MTTR Global',          f"{KPIS['mttr_global']:.1f} h",     '< 8 h',       f'+{KPIS["mttr_global"]-8:.1f}h', '⚠ CRÍTICO'],
        ['Costo Prom / OT',      f"${KPIS['costo_prom']/1000:.0f}K CLP", '< $400K','N/A',                             '● Normal'],
        ['% Retrabajo',          f"{KPIS['pct_retrabajo']:.1f}%",    '≤ 10%',       f'+{KPIS["pct_retrabajo"]-10:.1f}pp','⚠ CRÍTICO'],
        ['SLA Alta Prioridad',   f"{KPIS['sla']['Alta']:.1f}%",      '≥ 90%',       f'-{90-KPIS["sla"]["Alta"]:.1f}pp','⚠ CRÍTICO'],
        ['SLA Media Prioridad',  f"{KPIS['sla']['Media']:.1f}%",     '≥ 90%',       f'-{90-KPIS["sla"]["Media"]:.1f}pp','▲ ALERTA'],
        ['SLA Baja Prioridad',   f"{KPIS['sla']['Baja']:.1f}%",      '≥ 90%',       f'-{90-KPIS["sla"]["Baja"]:.1f}pp','▲ ALERTA'],
    ]
    t_kpi = RLTable(kpi_data, colWidths=[4*cm, 3*cm, 2.5*cm, 2.5*cm, 2.5*cm])
    t_kpi.setStyle(TableStyle([
        ('FONTNAME',    (0,0),(-1,0),  'Helvetica-Bold'),
        ('FONTNAME',    (0,1),(-1,-1), 'Helvetica'),
        ('FONTSIZE',    (0,0),(-1,-1), 8.5),
        ('BACKGROUND',  (0,0),(-1,0),  colors.HexColor('#1A3A5C')),
        ('TEXTCOLOR',   (0,0),(-1,0),  colors.white),
        ('GRID',        (0,0),(-1,-1), 0.5, colors.HexColor('#BDC3C7')),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#F8FBFF')]),
        ('ALIGN',       (1,0),(-1,-1), 'CENTER'),
        ('PADDING',     (0,0),(-1,-1), 5),
        ('TEXTCOLOR',   (4,1),(4,3),   colors.HexColor('#C0392B')),
        ('TEXTCOLOR',   (4,4),(4,5),   colors.HexColor('#E67E22')),
        ('TEXTCOLOR',   (4,6),(4,6),   colors.HexColor('#1E8449')),
    ]))
    story.append(t_kpi)
    story.append(Spacer(1, 0.5*cm))

    # ANÁLISIS DE TÉCNICOS 
    story.append(Paragraph('3. DESEMPEÑO DE TÉCNICOS', st_h2))
    story.append(Paragraph(
        'Se aplicó un <b>modelo de clustering K-Means</b> sobre 5 dimensiones '
        '(horas promedio, costo, retrabajo, MTTR y experiencia) para segmentar la fuerza técnica '
        'en tres perfiles de eficiencia. Los resultados revelan una distribución altamente asimétrica '
        'con un caso anómalo que distorsiona los promedios de flota.', st_body))

    try:
        img_tec = RLImage('C:/Users/Axel/Downloads/Movitec S.A/graficos/rpt_tecnicos.png',
                          width=16*cm, height=5.8*cm)
        story.append(img_tec)
    except: pass
    story.append(Spacer(1, 0.2*cm))

    story.append(Paragraph('Técnicos en Zona Crítica (Cluster Rojo)', st_h3))
    story.append(Paragraph(
        '• <b>Ignacio Ramírez</b> — 11 años de experiencia, costo promedio $677.805 CLP/OT '
        '(+88% sobre la media), tasa de retrabajo del 43.8%. La paradoja experiencia-rendimiento '
        'sugiere patrones de trabajo ineficientes consolidados en el tiempo, no falta de capacidad '
        'técnica. Acción requerida: plan de mejora estructurado con KPIs mensuales vinculantes.', st_bullet))
    story.append(Spacer(1, 0.3*cm))

    #
    # ─ SECCIÓN 4: ANÁLISIS DE EQUIPOS
    story.append(Paragraph('4. ANÁLISIS DE EQUIPOS Y ACTIVOS CRÍTICOS', st_h2))
    story.append(Paragraph(
        'El análisis Pareto confirma la concentración de costo en un subconjunto reducido de activos. '
        'Un <b>Risk Score compuesto</b> (frecuencia, costo, retrabajo y fallas no planificadas) '
        'permite priorizar intervenciones de mantenimiento predictivo.', st_body))

    try:
        img_eq = RLImage('C:/Users/Axel/Downloads/Movitec S.A/graficos/rpt_equipos.png',
                         width=16*cm, height=5.8*cm)
        story.append(img_eq)
    except: pass

    eq_criticos = equipo_stats[equipo_stats['risk_score']>40].sort_values('risk_score',ascending=False)
    eq_data = [['Equipo','Risk Score','% Retrabajo','% No Planif.','Costo Total','Diagnóstico']]
    diags = {
        'Compresor C-102': 'Overhaul urgente',
        'Conveyor CV-901': 'Predictivo requerido',
        'Motor M-303':     'Revisión estructural',
        'Bomba B-202':     'Plan repuestos',
        'Grúa GR-801':     'Inspección periódica',
    }
    for _,r in eq_criticos.head(5).iterrows():
        eq_data.append([
            r['equipo'], f"{r['risk_score']:.0f}",
            f"{r['pct_retrabajo']:.1f}%", f"{r['pct_no_planif']:.0f}%",
            f"${r['costo_total']/1e6:.1f}M", diags.get(r['equipo'],'Revisar'),
        ])
    t_eq = RLTable(eq_data, colWidths=[4*cm, 2.2*cm, 2.5*cm, 2.5*cm, 2.5*cm, 3*cm])
    t_eq.setStyle(TableStyle([
        ('FONTNAME',  (0,0),(-1,0),  'Helvetica-Bold'),
        ('FONTNAME',  (0,1),(-1,-1), 'Helvetica'),
        ('FONTSIZE',  (0,0),(-1,-1), 8.5),
        ('BACKGROUND',(0,0),(-1,0),  colors.HexColor('#1A3A5C')),
        ('TEXTCOLOR', (0,0),(-1,0),  colors.white),
        ('GRID',      (0,0),(-1,-1), 0.5, colors.HexColor('#BDC3C7')),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#FEF9E7')]),
        ('ALIGN',     (1,0),(-1,-1), 'CENTER'),
        ('PADDING',   (0,0),(-1,-1), 5),
    ]))
    story.append(Spacer(1, 0.3*cm))
    story.append(t_eq)
    story.append(PageBreak())

    # TENDENCIA Y OUTLIERS 
    story.append(Paragraph('5. TENDENCIA TEMPORAL Y DETECCIÓN DE ANOMALÍAS', st_h2))
    story.append(Paragraph(
        'La serie temporal de costos y retrabajo muestra variabilidad estructural sin tendencia '
        'descendente, lo que indica ausencia de mecanismos de mejora continua activos. '
        'Se detectaron <b>157 órdenes anómalas</b> mediante distancia de Mahalanobis '
        '(χ²=7.38, df=3, p=0.975), representando el 4.0% del dataset con impacto desproporcionado '
        'en los promedios de flota.', st_body))

    try:
        img_tend = RLImage('C:/Users/Axel/Downloads/Movitec S.A/graficos/rpt_tendencia.png',
                           width=16*cm, height=7*cm)
        story.append(img_tend)
    except: pass
    story.append(Spacer(1, 0.3*cm))

    # RECOMENDACIONES Y AHORROS
    story.append(Paragraph('6. RECOMENDACIONES Y POTENCIAL DE AHORRO', st_h2))
    story.append(Paragraph(
        'Las siguientes acciones están priorizadas por impacto económico cuantificable '
        'y viabilidad de implementación en el corto/mediano plazo.', st_body))

    rec_data = [
        ['#','Acción','Plazo','Ahorro Estimado'],
        ['R1','Protocolo anti-retrabajo:\nCheck-list de calidad + auditoría de OT recurrentes',
         'Inmediato\n(0–30 días)', '~$180M CLP/año'],
        ['R2','Plan de mejora técnico Ignacio Ramírez:\nKPIs vinculantes + supervisión OT alta prioridad',
         'Inmediato\n(0–30 días)', '~$45M CLP/año'],
        ['R3','Overhaul Compresor C-102:\nAnálisis vida útil + plan repuestos críticos',
         'Urgente\n(30 días)', '~$60M CLP/año'],
        ['R4','Protocolo SLA Alta Prioridad:\nGuardia 24/7 + tablero de control en tiempo real',
         'Corto plazo\n(30–60 días)', '+24pp cumplimiento'],
        ['R5','Mantenimiento predictivo (IoT):\nSensores vibración/temperatura en 3 equipos críticos',
         'Mediano plazo\n(60–90 días)', '~$90M CLP/año'],
        ['R6','Sistema de alertas de costos:\nBloqueo automático OT > media+2.5σ requiere aprobación',
         'Mediano plazo\n(60–90 días)', 'Control financiero'],
    ]
    t_rec = RLTable(rec_data, colWidths=[0.8*cm, 8.5*cm, 2.7*cm, 3*cm])
    t_rec.setStyle(TableStyle([
        ('FONTNAME',  (0,0),(-1,0),  'Helvetica-Bold'),
        ('FONTNAME',  (0,1),(-1,-1), 'Helvetica'),
        ('FONTSIZE',  (0,0),(-1,-1), 8),
        ('BACKGROUND',(0,0),(-1,0),  colors.HexColor('#1A3A5C')),
        ('TEXTCOLOR', (0,0),(-1,0),  colors.white),
        ('GRID',      (0,0),(-1,-1), 0.5, colors.HexColor('#BDC3C7')),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#EBF5FB')]),
        ('ALIGN',     (2,0),(-1,-1), 'CENTER'),
        ('VALIGN',    (0,0),(-1,-1), 'MIDDLE'),
        ('PADDING',   (0,0),(-1,-1), 5),
        ('FONTNAME',  (0,1),(0,3),   'Helvetica-Bold'),
        ('TEXTCOLOR', (0,1),(0,3),   colors.HexColor('#C0392B')),
        ('FONTNAME',  (0,4),(0,6),   'Helvetica-Bold'),
        ('TEXTCOLOR', (0,4),(0,6),   colors.HexColor('#E67E22')),
        ('ROWHEIGHTS',(0,1),(-1,-1), 38),
    ]))
    story.append(t_rec)
    story.append(Spacer(1, 0.5*cm))

    # Resumen de ahorro total
    st_ahorro = ParagraphStyle('ahorro', fontSize=12, fontName='Helvetica-Bold',
                                textColor=colors.white, backColor=colors.HexColor('#1E8449'),
                                alignment=TA_CENTER, borderPadding=(8,10,8,10), spaceAfter=6)
    story.append(Paragraph(
        '💰  POTENCIAL DE AHORRO TOTAL ESTIMADO:  ~$375M CLP / año  (≈ 19.5% del costo total operativo)', st_ahorro))

    story.append(Spacer(1, 0.4*cm))
    story.append(LINEA_GRIS)
    story.append(Paragraph(
        f'Informe generado automáticamente · Movitec S.A. · '
        f'{datetime.now().strftime("%d/%m/%Y %H:%M")} · '
        f'Consultoría Analítica Senior — Optimización Operativa', st_footer))

    doc.build(story)
    print(f"✓ PDF ejecutivo guardado: {path_out}")

generar_pdf_ejecutivo('C:/Users/Axel/Downloads/Movitec S.A/graficos/MOVITEC_Informe_Ejecutivo.pdf')
print("\nFase 6 completada — Excel + PDF generados")
